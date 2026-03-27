import streamlit as st
import pandas as pd
import re
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.title("BNI Transaction Database Generator")

uploaded_file = st.file_uploader(
    "Upload File (.xlsx / .xls / .csv)",
    type=["xlsx","xls","csv"]
)

# =====================================
# EXTRACT UNIQUE CODE BNI
# =====================================
def ambil_kode_unik(text):

    if pd.isna(text):
        return "N/A"

    text = str(text)
    lower = text.lower()

    if "otopay" in lower:
        return "IGNORE"

    m = re.search(r'PEMINDAHAN DARI\s+(\d+)', text)
    if m:
        return m.group(1)

    m = re.search(r'\|\s*(\d{16})', text)
    if m:
        return m.group(1)

    m = re.search(r'(\d{16})\s+[A-Za-z]', text)
    if m:
        return m.group(1)

    m = re.search(r'PENGIRIM\s+(.*)', text)
    if m:
        return m.group(1).strip()

    m = re.search(r'\|\s*\d+\s+[A-Z\s]+\s([A-Z\s]+)$', text)
    if m:
        return m.group(1).strip()

    if "pemindahan dari" in lower and "`" in text:
        return "N/A"

    if "jakom" in lower:
        return "IGNORE"

    return "N/A"

# =====================================
# DETECT HEADER
# =====================================
def detect_header(file):

    preview = pd.read_excel(file, header=None, nrows=20)

    for i in range(20):
        row = preview.iloc[i].astype(str).str.lower()
        if any("description" in cell for cell in row):
            return i

    return 0

# =====================================
# DETECT COLUMN
# =====================================
def detect_columns(df):

    df.columns = df.columns.str.strip()

    desc_col = None
    id_col = None

    for col in df.columns:
        if "description" in col.lower():
            desc_col = col
        if col.lower() == "id":
            id_col = col

    return id_col, desc_col

# =====================================
# MAIN PROCESS
# =====================================
if uploaded_file:

    try:

        if uploaded_file.name.endswith(".csv"):
            df = pd.read_csv(uploaded_file, sep=None, engine="python")
        else:
            header_row = detect_header(uploaded_file)
            df = pd.read_excel(uploaded_file, header=header_row)

        id_col, desc_col = detect_columns(df)

        if desc_col is None:
            st.error("Kolom Description tidak ditemukan")
            st.stop()

        if id_col is None:
            st.error("Kolom ID tidak ditemukan")
            st.stop()

        df["KODE_UNIK"] = df[desc_col].apply(ambil_kode_unik)

        database = df[[id_col, "KODE_UNIK", desc_col]].copy()

        database = database.rename(columns={
            id_col: "ID",
            desc_col: "Description"
        })

        database["ID"] = pd.to_numeric(database["ID"], errors="coerce")

        database = database[database["KODE_UNIK"] != "IGNORE"]

        valid = database[database["KODE_UNIK"] != "N/A"].copy()
        anomali = database[database["KODE_UNIK"] == "N/A"].copy()

        # =====================================
        # GROUPING LOGIC (INI YANG BARU)
        # =====================================
        grouped = []

        used_index = set()

        for i, row in valid.iterrows():
            if i in used_index:
                continue

            same_group = valid[
                (valid["KODE_UNIK"] == row["KODE_UNIK"]) |
                (valid["ID"] == row["ID"])
            ]

            used_index.update(same_group.index)

            ids = sorted(set(same_group["ID"].dropna()))
            kode = sorted(set(same_group["KODE_UNIK"]))
            desc = list(same_group["Description"])

            grouped.append({
                "ID": " ; ".join(map(lambda x: str(int(x)), ids)),
                "KODE_UNIK": " ; ".join(kode),
                "Description": " ; ".join(desc),
                "TYPE": "DOUBLE" if len(same_group) > 1 else "NORMAL",
                "SORT_KEY": min(ids) if len(ids) > 0 else 0
            })

        grouped_df = pd.DataFrame(grouped)

        normal = grouped_df[grouped_df["TYPE"] == "NORMAL"].copy()
        double = grouped_df[grouped_df["TYPE"] == "DOUBLE"].copy()

        normal = normal.sort_values("SORT_KEY")
        double = double.sort_values("SORT_KEY")
        anomali = anomali.sort_values("ID")

        anomali["TYPE"] = "NA"

        hasil = pd.concat([
            normal.drop(columns=["TYPE","SORT_KEY"]),
            double.drop(columns=["TYPE","SORT_KEY"]),
            anomali
        ], ignore_index=True)

        # =====================================
        # DASHBOARD
        # =====================================
        col1, col2, col3 = st.columns(3)

        col1.metric("Total transaksi", len(database))
        col2.metric("Database bersih", len(normal) + len(double))
        col3.metric("Perlu cek manual (N/A)", len(anomali))

        st.success("Database berhasil dibuat")
        st.dataframe(hasil)

        # =====================================
        # EXPORT + COLORING
        # =====================================
        output = BytesIO()
        hasil.to_excel(output, index=False)

        output.seek(0)
        wb = load_workbook(output)
        ws = wb.active

        yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        for i, row in enumerate(grouped):
            excel_row = i + 2

            if row["TYPE"] == "DOUBLE":
                for col in range(1, 4):
                    ws.cell(row=excel_row, column=col).fill = yellow

        start_na = len(normal) + len(double) + 2

        for row in range(start_na, start_na + len(anomali)):
            for col in range(1, 4):
                ws.cell(row=row, column=col).fill = red

        final_output = BytesIO()
        wb.save(final_output)

        st.download_button(
            "Download DATABASE_HASIL_BNI.xlsx",
            final_output.getvalue(),
            "DATABASE_HASIL_BNI.xlsx"
        )

    except Exception as e:
        st.error("Terjadi error saat memproses file")
        st.write(e)
